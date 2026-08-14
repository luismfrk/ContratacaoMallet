from __future__ import annotations

from copy import copy
from io import BytesIO
from pathlib import Path
import re
import unicodedata

from openpyxl import load_workbook
import pdfplumber


BASE_DIR = Path(__file__).resolve().parent
MODELOS = {
    "material": BASE_DIR / "modelo_material.xlsx",
    "servico": BASE_DIR / "modelo_servico.xlsx",
}


def _texto(valor) -> str:
    return "" if valor is None else str(valor).strip()


def _normalizar(valor) -> str:
    texto = unicodedata.normalize("NFKD", _texto(valor)).encode("ascii", "ignore").decode()
    return re.sub(r"\s+", " ", texto.lower())


def _numero(valor, padrao=0.0) -> float:
    if isinstance(valor, (int, float)):
        return float(valor)
    texto = _texto(valor).replace("R$", "").replace("%", "").replace(" ", "")
    if not texto:
        return padrao
    if "," in texto:
        texto = texto.replace(".", "").replace(",", ".")
    try:
        return float(texto)
    except ValueError:
        return padrao


def ler_orcamento_excel(conteudo: bytes) -> dict:
    try:
        workbook = load_workbook(BytesIO(conteudo), data_only=True)
    except Exception as exc:
        raise ValueError("O orçamento deve ser uma planilha Excel .xlsx válida.") from exc

    melhor = None
    for planilha in workbook.worksheets:
        for linha in range(1, min(planilha.max_row, 100) + 1):
            cabecalhos = {_normalizar(planilha.cell(linha, coluna).value): coluna for coluna in range(1, planilha.max_column + 1)}
            descricao = next((c for h, c in cabecalhos.items() if any(x in h for x in ("descricao", "produto", "servico"))), None)
            if not descricao:
                descricao = next((c for h, c in cabecalhos.items() if h == "item"), None)
            quantidade = next((c for h, c in cabecalhos.items() if any(x in h for x in ("quantidade", "qtd", "qtde"))), None)
            unitario = next((c for h, c in cabecalhos.items() if "unit" in h or "preco" in h or "valor" in h), None)
            if descricao and (quantidade or unitario):
                melhor = (planilha, linha, descricao, quantidade, unitario)
                break
        if melhor:
            break
    if not melhor:
        raise ValueError("Não encontrei colunas de descrição, quantidade e valor unitário no orçamento.")

    planilha, cabecalho, descricao, quantidade, unitario = melhor
    itens = []
    vazias = 0
    for linha in range(cabecalho + 1, planilha.max_row + 1):
        nome = _texto(planilha.cell(linha, descricao).value)
        if not nome:
            vazias += 1
            if vazias >= 5 and itens:
                break
            continue
        vazias = 0
        normalizado = _normalizar(nome)
        if normalizado.startswith("total") or normalizado in {"subtotal", "valor total"}:
            break
        qtd = _numero(planilha.cell(linha, quantidade).value, 1) if quantidade else 1
        valor = _numero(planilha.cell(linha, unitario).value) if unitario else 0
        if qtd > 0 and (valor > 0 or nome):
            itens.append({"descricao": nome, "quantidade": qtd, "valor_unitario": valor})
    if not itens:
        raise ValueError("A planilha foi lida, mas nenhum item de orçamento foi encontrado.")
    if len(itens) > 29:
        raise ValueError("O modelo comporta até 29 itens por requisição.")
    return {"arquivo": planilha.title, "itens": itens, "total": sum(i["quantidade"] * i["valor_unitario"] for i in itens)}


def _achar_colunas(cabecalhos: list) -> tuple[int | None, int | None, int | None, int | None, int | None]:
    nomes = [_normalizar(valor) for valor in cabecalhos]
    descricao = next((i for i, h in enumerate(nomes) if any(x in h for x in ("descricao", "produto", "servico", "especificacao"))), None)
    quantidade = next((i for i, h in enumerate(nomes) if any(x in h for x in ("quantidade", "qtd", "qtde", "quant."))), None)
    unitario = next((i for i, h in enumerate(nomes) if "unit" in h or "preco" in h), None)
    total = next((i for i, h in enumerate(nomes) if "total" in h), None)
    desconto = next((i for i, h in enumerate(nomes) if "% a/d" in h or "desconto" in h or h in {"a/d", "%ad"}), None)
    return descricao, quantidade, unitario, total, desconto


def ler_orcamento_pdf(conteudo: bytes) -> dict:
    try:
        pdf = pdfplumber.open(BytesIO(conteudo))
    except Exception as exc:
        raise ValueError("O arquivo não é um PDF válido.") from exc
    itens = []
    metadados = {}
    tipos_encontrados = set()
    tinha_texto = False
    with pdf:
        for pagina in pdf.pages:
            texto_pagina = pagina.extract_text() or ""
            tinha_texto = tinha_texto or bool(texto_pagina.strip())
            linhas_texto = texto_pagina.splitlines()
            if not metadados.get("fornecedor") and linhas_texto:
                metadados["fornecedor"] = linhas_texto[0].strip()
            cnpj_indice = next((i for i, linha in enumerate(linhas_texto[:8]) if "CNPJ:" in linha.upper()), None)
            if cnpj_indice is not None and cnpj_indice + 1 < len(linhas_texto):
                metadados.setdefault("endereco", linhas_texto[cnpj_indice + 1].strip())
            placa = re.search(r"PLACA:\s*([A-Z0-9-]+)", texto_pagina, re.I)
            orcamento = re.search(r"(?:N[°º]\s*)?OR[ÇC]AMENTO:\s*([\w/-]+)", texto_pagina, re.I)
            identificadores = []
            if placa:
                identificadores.append(f"PLACA- {placa.group(1)}")
                metadados["placa"] = placa.group(1)
            if orcamento:
                identificadores.append(f"ORÇAMENTO- {orcamento.group(1)}")
                metadados["numero_orcamento"] = orcamento.group(1)
            if identificadores:
                metadados["identificacao"] = "   ".join(identificadores)
            cidade = re.search(r"CEP[^\n]*?\s[-–]\s*([A-ZÀ-Ý ]+?)(?:-PR|/PR|\n)", texto_pagina, re.I)
            if cidade:
                metadados.setdefault("cidade", cidade.group(1).strip())
            for tabela in pagina.extract_tables() or []:
                colunas = None
                secao = ""
                for linha in tabela:
                    primeira = _normalizar(linha[0] if linha else "")
                    if "mao de obra" in primeira or "servico" in primeira:
                        secao = "servico"
                    elif "pecas" in primeira or "produtos" in primeira:
                        secao = "material"
                    achadas = _achar_colunas(linha or [])
                    if achadas[0] is not None and (achadas[1] is not None or achadas[2] is not None):
                        colunas = achadas
                        continue
                    if colunas is None:
                        continue
                    descricao, quantidade, unitario, total, desconto_coluna = colunas
                    if not linha or descricao >= len(linha):
                        continue
                    nome = _texto(linha[descricao]).replace("\n", " ")
                    nome_normalizado = _normalizar(nome)
                    if not nome or nome_normalizado.startswith(("total", "subtotal")):
                        continue
                    qtd = _numero(linha[quantidade], 1) if quantidade is not None and quantidade < len(linha) else 1
                    valor = _numero(linha[unitario]) if unitario is not None and unitario < len(linha) else 0
                    if not valor and total is not None and total < len(linha) and qtd:
                        valor = _numero(linha[total]) / qtd
                    if qtd > 0 and valor > 0:
                        desconto = _numero(linha[desconto_coluna]) if desconto_coluna is not None and desconto_coluna < len(linha) else 0
                        itens.append({"descricao": nome, "quantidade": qtd, "valor_unitario": round(valor, 2), "desconto": desconto, "tipo": secao or "material"})
                        tipos_encontrados.add(secao or "material")
    if not itens:
        if not tinha_texto:
            raise ValueError("Este PDF parece digitalizado como imagem. Envie um PDF com texto pesquisável ou aplique OCR antes de importar.")
        raise ValueError("Não encontrei no PDF uma tabela com descrição, quantidade e valor unitário.")
    if len(itens) > 29:
        raise ValueError("O modelo comporta até 29 itens por requisição.")
    tipo = "servico" if tipos_encontrados == {"servico"} else "material"
    grupos = {
        nome: {
            "itens": [item for item in itens if item["tipo"] == nome],
            "total": sum(item["quantidade"] * item["valor_unitario"] for item in itens if item["tipo"] == nome),
        }
        for nome in ("material", "servico") if any(item["tipo"] == nome for item in itens)
    }
    return {"arquivo": "PDF", "itens": itens, "total": sum(i["quantidade"] * i["valor_unitario"] for i in itens), "tipo": tipo, "misto": len(grupos) > 1, "grupos": grupos, "metadados": metadados}


def ler_orcamento(conteudo: bytes, nome_arquivo: str = "") -> dict:
    if nome_arquivo.lower().endswith(".pdf") or conteudo[:4] == b"%PDF":
        return ler_orcamento_pdf(conteudo)
    return ler_orcamento_excel(conteudo)


def gerar_requisicao(tipo: str, dados: dict) -> tuple[bytes, str]:
    if tipo not in MODELOS:
        raise ValueError("Tipo de requisição inválido.")
    itens = dados.get("itens") or []
    if not itens or len(itens) > 29:
        raise ValueError("Informe de 1 a 29 itens.")
    workbook = load_workbook(MODELOS[tipo])
    ws = workbook.active
    fornecedor = _texto(dados.get("fornecedor"))
    endereco = _texto(dados.get("endereco"))
    cidade = _texto(dados.get("cidade"))
    destino = _texto(dados.get("destino")) or "SEC. EDUCAÇÃO - SETOR TRANSPORTE"
    fonte = _texto(dados.get("fonte_recurso"))
    identificacao = _texto(dados.get("identificacao"))
    desconto = _numero(dados.get("desconto"))
    if desconto > 1:
        desconto /= 100

    ws["B11"] = f"Solicitamos que seja emitida Autorização de Fornecimento dos materiais/serviços abaixo descritos destinados a:      {destino}"
    ws["B13"] = f"Fornecedor: {fornecedor}"
    ws["B15"] = f"Endereço: {endereco}"
    ws["B17"] = f"Cidade: {cidade}"
    ws["B19"] = f"Fonte/Recurso: {fonte}                              {identificacao}"
    ws["E21"] = "R$ (Hora)" if tipo == "servico" else "R$ (Unitário)"
    for linha in range(23, 52):
        for coluna in (2, 4, 5, 6, 7, 8):
            ws.cell(linha, coluna).value = None
    for indice, item in enumerate(itens, 23):
        qtd = _numero(item.get("quantidade"), 1)
        valor = _numero(item.get("valor_unitario"))
        ws.cell(indice, 2).value = qtd
        ws.cell(indice, 4).value = _texto(item.get("descricao"))
        ws.cell(indice, 5).value = valor
        desconto_item = _numero(item.get("desconto"), desconto)
        if desconto_item > 1:
            desconto_item /= 100
        ws.cell(indice, 6).value = desconto_item
        ws.cell(indice, 7).value = f"=E{indice}-(E{indice}*F{indice})"
        ws.cell(indice, 8).value = f"=G{indice}*B{indice}"
    ws["H52"] = "=SUM(H23:H51)"
    saida = BytesIO()
    workbook.save(saida)
    placa = _texto(dados.get("placa"))
    numero_orcamento = _texto(dados.get("numero_orcamento"))
    if placa and numero_orcamento:
        categoria = "PEÇAS" if tipo == "material" else "SERVIÇO"
        nome_arquivo = f"{placa} {numero_orcamento} {categoria}.xlsx"
    else:
        nome = re.sub(r"[^A-Za-z0-9_-]+", "_", identificacao or tipo).strip("_")
        nome_arquivo = f"requisicao_{nome}.xlsx"
    return saida.getvalue(), nome_arquivo
