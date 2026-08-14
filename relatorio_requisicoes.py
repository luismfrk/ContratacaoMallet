from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

MESES = ("", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
         "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro")


def gerar_relatorio_requisicoes(linhas: list[dict], ano: int, mes: int, secretaria: str) -> tuple[bytes, str]:
    wb = Workbook()
    ws = wb.active
    ws.title = "Gastos por veículo"
    ws.merge_cells("A1:E1")
    ws["A1"] = "RELATÓRIO MENSAL DE GASTOS POR VEÍCULO"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="0873B9")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:E2")
    ws["A2"] = f"{MESES[mes]} de {ano}" + (f" — {secretaria}" if secretaria else " — Todas as secretarias")
    ws["A2"].alignment = Alignment(horizontal="center")
    cabecalhos = ("Veículo / Placa", "Requisições", "Peças", "Serviços", "Total gasto")
    for coluna, titulo in enumerate(cabecalhos, 1):
        celula = ws.cell(4, coluna, titulo)
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = PatternFill("solid", fgColor="075D9C")
    for indice, linha in enumerate(linhas, 5):
        valores = (linha["placa"], linha["requisicoes"], float(linha["materiais"] or 0),
                   float(linha["servicos"] or 0), float(linha["total"] or 0))
        for coluna, valor in enumerate(valores, 1):
            ws.cell(indice, coluna, valor)
        for coluna in range(3, 6):
            ws.cell(indice, coluna).number_format = 'R$ #,##0.00'
    total_linha = 5 + len(linhas)
    ws.cell(total_linha, 1, "TOTAL DO MÊS").font = Font(bold=True)
    for coluna in range(3, 6):
        letra = get_column_letter(coluna)
        ws.cell(total_linha, coluna, f"=SUM({letra}5:{letra}{total_linha - 1})")
        ws.cell(total_linha, coluna).font = Font(bold=True)
        ws.cell(total_linha, coluna).number_format = 'R$ #,##0.00'
    for coluna, largura in enumerate((22, 14, 18, 18, 20), 1):
        ws.column_dimensions[get_column_letter(coluna)].width = largura
    ws.freeze_panes = "A5"
    saida = BytesIO()
    wb.save(saida)
    return saida.getvalue(), f"RELATORIO VEICULOS {ano}-{mes:02d}.xlsx"
