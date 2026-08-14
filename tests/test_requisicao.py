from io import BytesIO

from openpyxl import Workbook, load_workbook

from requisicao import gerar_requisicao, ler_orcamento


def _orcamento() -> bytes:
    workbook = Workbook()
    planilha = workbook.active
    planilha.append(["ITEM", "DESCRIÇÃO", "QTD", "VALOR UNITÁRIO"])
    planilha.append([1, "Filtro de óleo", 2, 35.5])
    planilha.append([2, "Correia", 1, 90])
    arquivo = BytesIO()
    workbook.save(arquivo)
    return arquivo.getvalue()


def test_le_orcamento_por_cabecalhos():
    resultado = ler_orcamento(_orcamento())
    assert resultado["itens"][0] == {
        "descricao": "Filtro de óleo",
        "quantidade": 2.0,
        "valor_unitario": 35.5,
    }
    assert resultado["total"] == 161


def test_gera_requisicao_com_formulas_e_desconto():
    itens = ler_orcamento(_orcamento())["itens"]
    conteudo, nome = gerar_requisicao(
        "material",
        {
            "fornecedor": "Fornecedor Teste",
            "endereco": "Rua A",
            "cidade": "Mallet",
            "fonte_recurso": "1000",
            "identificacao": "ORDEM 1/2026",
            "desconto": 15,
            "itens": itens,
        },
    )
    planilha = load_workbook(BytesIO(conteudo), data_only=False).active
    assert nome == "requisicao_ORDEM_1_2026.xlsx"
    assert planilha["D23"].value == "Filtro de óleo"
    assert planilha["F23"].value == 0.15
    assert planilha["H23"].value == "=G23*B23"
    assert planilha["H52"].value == "=SUM(H23:H51)"
